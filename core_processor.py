import os
import re
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.styles import Border

# --- CONFIGURACIÓN GLOBAL (Cerrado a modificación, abierto a expansión) ---
NOMINA_MAPPER = {
    "Empleado Fijo Administrativo": "EF ADM",
    "Empleado Fijo Enfermeros": "EF ENFER",
    "Empleado Fijo Médicos": "EF MEDICO",
    "Empleado Fijo Otros Gremios": "EF OTROS GRE",
    "Obrero Fijo": "OF OBRERO",
    "Empleado Contratado Administrativo": "EC ADM",
    "Empleado Contratado Enfermeros": "EC ENFER",
    "Empleado Contratado Médicos": "EC MEDICO",
    "Empleado Contratado Otros Gremios": "EC OTROS GRE",
    "Obrero Contratado": "OC OBRERO"
}

VALID_HEADER = "Personal Con Cuenta Bancaria Activa"

# --- CLASES ---

class PDFWorker:
    """S de SOLID: Responsabilidad única - Extraer datos de PDF."""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)

    def process(self):
        """Retorna un dict con los datos si es válido, sino None."""
        try:
            doc = fitz.open(self.filepath)
            text = "".join([page.get_text() for page in doc])
            
            # Validación de seguridad
            if VALID_HEADER not in text:
                return None

            # Extracción de Grupo de Nómina
            grupo_match = re.search(r"Grupo de Nómina:\s*(.*)", text)
            grupo_raw = grupo_match.group(1).strip() if grupo_match else None
            
            # Traducción mediante el mapper
            gremio_excel = NOMINA_MAPPER.get(grupo_raw)
            
            if not gremio_excel:
                return None

            # Extracción de Totales (buscamos la última ocurrencia)
            monto_matches = re.findall(r"Total General Bs\.:\s*([\d\.,]+)", text)
            trab_matches = re.findall(r"Total Trabajadores:\s*(\d+)", text)
            
            return {
                "gremio": gremio_excel,
                "monto": float(monto_matches[-1].replace('.', '').replace(',', '.')) if monto_matches else 0.0,
                "trabajadores": int(trab_matches[-1]) if trab_matches else 0
            }
        except Exception as e:
            print(f"Error procesando {self.filename}: {e}")
            return None

class ExcelBox:
    """O de SOLID: Maneja la lógica de un cuadro. Se instancia por sección."""
    
    def __init__(self, sheet, anchor_name):
        self.sheet = sheet
        self.anchor_name = anchor_name
        self.anchor_cell = self._find_anchor()
        self.data_rows = self._map_box_structure()

    def _find_anchor(self):
        """Busca la celda que dice '1ERA QCNA', '2DA QCNA', etc."""
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == self.anchor_name:
                    return cell
        return None

    def _map_box_structure(self):
        """
        Algoritmo de bordes: Identifica qué filas pertenecen al cuadro
        basándose en si tienen bordes y el nombre del gremio.
        """
        mapping = {}
        if not self.anchor_cell:
            return mapping

        # Empezamos la búsqueda desde la columna de 'GREMIOS' 
        # (Basado en tu wireframe, el gremio suele estar a la derecha del anclaje o en la misma fila)
        start_row = self.anchor_cell.row
        
        # Escaneamos hacia abajo hasta que no haya más bordes
        for row_idx in range(start_row, start_row + 20): # Límite razonable de 20 filas por cuadro
            # Buscamos en la columna de GREMIOS (ejemplo: columna D es la 4)
            # Aquí ajustaremos según el wireframe real
            for col_idx in range(1, 15): 
                cell = self.sheet.cell(row=row_idx, column=col_idx)
                if cell.value in NOMINA_MAPPER.values():
                    # Si tiene borde, es parte de nuestro cuadro
                    if cell.border.left.style or cell.border.bottom.style:
                        mapping[cell.value] = row_idx
        return mapping

    def fill(self, gremio, monto, cant):
        """Escribe los datos en las columnas de Cantidad y Listado Banco."""
        if gremio in self.data_rows:
            target_row = self.data_rows[gremio]
            # TODO: Localizar columnas exactas de 'CANT. DE TRABAJADORES' y 'LISTADOS DE BANCO'
            # Por ahora, un placeholder:
            print(f"Escribiendo en {self.anchor_name} -> {gremio}: {cant} trab, {monto} Bs.")