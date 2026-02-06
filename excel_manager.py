from openpyxl.utils import get_column_letter

class ExcelBox:
    # Parámetros de expansión para delimitar el área de búsqueda desde el "ancla"
    FILAS_ARRIBA = 5
    FILAS_ABAJO  = 15
    COL_DERECHA  = 9

    def __init__(self, sheet, anchor_name, mapper):
        self.sheet = sheet
        self.anchor_name = anchor_name
        self.mapper = mapper
        
        self.col_indices = {}  # Almacenará {'monto': index, 'cant': index, 'gremios': index}
        self.data_rows = {}    # Almacenará {'EF ADM': row_index, ...}
        
        # Proceso de inicialización
        self.anchor_cell = self._find_anchor()
        if self.anchor_cell:
            self._initialize_structure()

    def _find_anchor(self):
        """Busca la celda de origen (ej. '1ERA QCNA') en toda la hoja."""
        for row in self.sheet.iter_rows():
            for cell in row:
                if str(cell.value).strip() == self.anchor_name:
                    return cell
        return None

    def _get_value_safe(self, cell):
        """Obtiene el valor de una celda, manejando correctamente las celdas combinadas."""
        val = cell.value
        if val is None:
            for merged_range in self.sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # En openpyxl, el valor de una celda combinada reside en la celda superior izquierda
                    val = merged_range.start_cell.value
                    break
        return val

    def _map_columns(self, r_range, c_range):
        """Busca las cabeceras normalizando puntos y comas, dentro del área definida."""
        for r in r_range:
            for c in c_range:
                cell = self.sheet.cell(row=r, column=c)
                # Normalización: quitamos puntos y comas para comparar texto limpio
                val = str(cell.value or "").upper().replace(".", "").replace(",", "").strip()
                
                if "GREMIOS" in val:
                    self.col_indices['gremios'] = c
                elif "LISTADOS DE BANCO" in val:
                    self.col_indices['monto'] = c
                elif "CANT DE TRABAJADORES" in val and "INACTIVO" not in val:
                    self.col_indices['cant'] = c
                # Nuevas columnas para inactivos
                elif "MONTO TRABAJADORES (INACTIVOS)" in val or "MONTO TRABAJADORES INACTIVOS" in val:
                    self.col_indices['monto_inactivo'] = c
                elif "CANT DE TRABAJ INACTIVO" in val or "CANTIDAD DE TRABAJADORES INACTIVOS" in val:
                    self.col_indices['cant_inactivo'] = c
                # --- AQUÍ ESTÁ EL CAMBIO PARA INACTIVOS ---
                # Usamos "IN" para capturar "INACTIVO" o "(INACTIVOS)"
                elif "MONTO" in val and "INACT" in val:
                    self.col_indices['monto_inactivo'] = c
                elif "CANT" in val and "INACT" in val:
                    self.col_indices['cant_inactivo'] = c
    def _map_rows(self, r_range):
        """Mapea las filas correspondientes a cada sigla (gremio) del mapper."""
        if 'gremios' not in self.col_indices:
            return

        col_g = self.col_indices['gremios']
        siglas_validas = [v[0] for v in self.mapper.values()]
        
        for r in r_range:
            cell = self.sheet.cell(row=r, column=col_g)
            val = self._get_value_safe(cell)
            val_str = str(val or "").strip()
            
            if val_str in siglas_validas:
                self.data_rows[val_str] = r

    def _initialize_structure(self):
        """Orquestador del mapeo de la estructura del cuadro en el Excel."""
        r_orig = self.anchor_cell.row
        c_orig = self.anchor_cell.column
        
        # Definir rangos de búsqueda
        filas_busqueda = range(r_orig - self.FILAS_ARRIBA, r_orig + self.FILAS_ABAJO + 1)
        cols_busqueda = range(c_orig, c_orig + self.COL_DERECHA + 1)

        # Ejecutar mapeos atómicos
        self._map_columns(filas_busqueda, cols_busqueda)
        self._map_rows(filas_busqueda)

        # Logs de control
        print(f"📍 Estructura '{self.anchor_name}' mapeada en {get_column_letter(c_orig)}{r_orig}")
        print(f"✅ Columnas halladas: {list(self.col_indices.keys())}")
        print(f"✅ Filas de gremios vinculadas: {len(self.data_rows)}")

    def fill(self, sigla, monto, cant, estado="ACTIVO"):
        """Escribe en las columnas correspondientes según el estado."""
        if sigla not in self.data_rows: return
        
        row = self.data_rows[sigla]
        
        # Lógica de decisión de columna
        if estado == "INACTIVO":
            # Si es inactivo, intentamos usar las columnas de inactivos
            col_m = self.col_indices.get('monto_inactivo')
            col_c = self.col_indices.get('cant_inactivo')
            
            # Si por alguna razón no mapeó las de inactivos, avisamos
            if not col_m or not col_c:
                print(f"⚠️ [!] No encontré columnas de INACTIVOS para {sigla} en {self.anchor_name}")
                return
        else:
            # Si es activo, usamos las normales
            col_m = self.col_indices.get('monto')
            col_c = self.col_indices.get('cant')

        if col_m: self.sheet.cell(row=row, column=col_m).value = monto
        if col_c: self.sheet.cell(row=row, column=col_c).value = cant
        
        print(f"✍️ [{estado}] {sigla} -> Monto: {monto} | Cant: {cant} en fila {row}")